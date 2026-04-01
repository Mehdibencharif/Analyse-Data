import streamlit as st
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from io import BytesIO
from datetime import timedelta
import re
import hashlib
import unicodedata
from pandas.api.types import is_numeric_dtype



# ----------------------------- Utilitaires "qualité data" -----------------------------



def coerce_numeric_general(df: pd.DataFrame, threshold: float = 0.6) -> pd.DataFrame:
    for col in df.columns:
        if col.lower() in ("timestamp", "notes"):
            continue
        s = df[col]
        if not is_numeric_dtype(s):
            s2 = s.astype(str).str.replace(",", ".", regex=False).str.strip()
            s2 = s2.replace(list(PLACEHOLDER_NULLS), pd.NA)
            numeric = pd.to_numeric(s2, errors="coerce")
            if numeric.notna().mean() >= threshold:
                df[col] = numeric
    return df



PLACEHOLDER_NULLS = {"", " ", "-", "—", "–", "NA", "N/A", "na", "n/a", "null", "None"}



def series_with_true_nans(s: pd.Series) -> pd.Series:
    if s.dtype == object:
        s = s.astype(str).str.strip()
        s = s.replace(list(PLACEHOLDER_NULLS), pd.NA)
        s = s.replace(r"^\s+$", pd.NA, regex=True)
    return s



TEMP_NAME_RE = re.compile(r"(?i)(temp|temperature|°\s*c|degc|degre|°c|\[°c\])")



def coerce_temperature_columns(df: pd.DataFrame) -> pd.DataFrame:
    for col in df.columns:
        if col.lower() in ("timestamp", "notes"):
            continue
        name = str(col)
        if TEMP_NAME_RE.search(name):
            s = df[col]
            if s.dtype == object:
                s = s.astype(str).str.replace(",", ".", regex=False).str.strip()
                s = s.replace(list(PLACEHOLDER_NULLS), pd.NA)
            df[col] = pd.to_numeric(s, errors="coerce")
    return df



# ----------------------------- Statistiques descriptives -----------------------------



def calculer_stats_descriptives(df: pd.DataFrame) -> pd.DataFrame:
    """
    Calcule min, max, moyenne, médiane et écart-type pour chaque colonne numérique.
    Retourne un DataFrame avec une ligne par capteur.
    """
    rows = []
    for col in df.columns:
        if str(col).lower() in ("timestamp", "notes"):
            continue
        s = series_with_true_nans(df[col])
        s_num = pd.to_numeric(s, errors="coerce").dropna()
        if len(s_num) > 0:
            rows.append({
                "Capteur": str(col).strip(),
                "Min": round(s_num.min(), 4),
                "Max": round(s_num.max(), 4),
                "Moyenne": round(s_num.mean(), 4),
                "Médiane": round(s_num.median(), 4),
                "Écart-type": round(s_num.std(), 4),
                "N valeurs": int(len(s_num)),
            })
        else:
            rows.append({
                "Capteur": str(col).strip(),
                "Min": None,
                "Max": None,
                "Moyenne": None,
                "Médiane": None,
                "Écart-type": None,
                "N valeurs": 0,
            })
    return pd.DataFrame(rows)



# ----------------------------- Streamlit : page & paramètres -----------------------------



st.set_page_config(page_title="Analyse de données capteurs", layout="wide")
st.title("📊 Analyse de données ")



st.sidebar.header("Paramètres d'analyse")
frequence = st.sidebar.selectbox(
    "Choisissez la fréquence d'analyse :",
    ["1min", "2min", "5min", "10min", "15min", "1H"]
)
rule_map = {
    "2min": "2min",
    "1min": "1min",
    "5min": "5min",
    "10min": "10min",
    "15min": "15min",
    "1H": "1H"
}



st.sidebar.subheader("Téléversement des fichiers")
main_file = st.sidebar.file_uploader(
    "📂 Fichier principal (obligatoire)",
    type=[".xlsx", ".xls", ".xlsm"],
    key="main"
)
compare_file = st.sidebar.file_uploader(
    "📂 Fichier de comparaison (facultatif)",
    type=[".xlsx", ".xls", ".xlsm"],
    key="compare"
)



def file_sha1(uploaded):
    data = uploaded.getvalue() if uploaded is not None else b""
    return hashlib.sha1(data).hexdigest()[:10]



if main_file:
    st.sidebar.caption(f"Hash fichier principal : `{file_sha1(main_file)}`")
if compare_file:
    st.sidebar.caption(f"Hash fichier comparaison : `{file_sha1(compare_file)}`")



if "last_main_sha1" not in st.session_state:
    st.session_state.last_main_sha1 = None
curr_sha1 = file_sha1(main_file) if main_file else None
if curr_sha1 and curr_sha1 != st.session_state.last_main_sha1:
    for k in list(st.session_state.keys()):
        if str(k).startswith("Fichier principal_sheet_"):
            del st.session_state[k]
    st.session_state.last_main_sha1 = curr_sha1



# ----------------------------- Chargement fichier -----------------------------



def charger_et_resampler(fichier, nom_fichier):
    raw = fichier.getvalue()
    xls = pd.ExcelFile(BytesIO(raw))
    sheet_key = f"{nom_fichier}_sheet_{hashlib.sha1(raw).hexdigest()[:8]}"

    feuille = xls.sheet_names[0] if len(xls.sheet_names) == 1 else st.selectbox(
        f"📄 Feuille à utiliser pour {nom_fichier}",
        xls.sheet_names,
        key=sheet_key
    )
    df = pd.read_excel(xls, sheet_name=feuille)
    df.columns = [str(c).strip() for c in df.columns]
    df = df.rename(columns={df.columns[0]: "timestamp"})
    df["timestamp"] = pd.to_datetime(df["timestamp"], errors="coerce")
    df = df.dropna(subset=["timestamp"]).sort_values("timestamp").reset_index(drop=True)
    return df



if not main_file:
    st.warning("⚠️ Veuillez téléverser un fichier principal pour démarrer l'analyse.")
    st.stop()



st.sidebar.subheader("Filtre temporel (optionnel)")
date_deb = st.sidebar.date_input("Début", value=None)
date_fin = st.sidebar.date_input("Fin", value=None)



def filtrer_periode(df):
    if date_deb:
        df = df[df["timestamp"] >= pd.Timestamp(date_deb)]
    if date_fin:
        df = df[df["timestamp"] <= pd.Timestamp(date_fin) + pd.Timedelta(days=1) - pd.Timedelta(microseconds=1)]
    return df



df_main = charger_et_resampler(main_file, "Fichier principal")
df_main = coerce_temperature_columns(df_main)
df_main = coerce_numeric_general(df_main)
df_main = filtrer_periode(df_main)



if not df_main.empty:
    st.sidebar.caption(f"Période détectée : {df_main['timestamp'].min()} → {df_main['timestamp'].max()}")



# ----------------------------- Nettoyage noms (comparaison) -----------------------------



def nettoyer_nom_capteur(nom: str) -> str:
    s = str(nom)
    s = re.sub(r"\s*[\[\(].*?[\]\)]", "", s)
    return s.strip()



df_main_cleaned = df_main.copy()
df_main_cleaned.columns = [
    "timestamp" if c == "timestamp" else nettoyer_nom_capteur(c)
    for c in df_main_cleaned.columns
]



# ----------------------------- Fichier de référence (facultatif) -----------------------------



df_compare = None
capteurs_reference = None
capteurs_reference_cleaned = None



if compare_file:
    try:
        df_compare = pd.read_excel(BytesIO(compare_file.getvalue()))
        if "Description" not in df_compare.columns:
            st.error("❌ Le fichier de comparaison doit contenir une colonne 'Description'.")
            st.stop()

        df_compare["Description"] = df_compare["Description"].astype(str).str.strip()
        capteurs_reference = set(df_compare["Description"])
        capteurs_reference_cleaned = {nettoyer_nom_capteur(c) for c in capteurs_reference}

        st.success("✅ Fichier de comparaison chargé avec succès.")
    except Exception as e:
        st.error(f"❌ Erreur lors de la lecture du fichier de comparaison : {e}")
        st.stop()
else:
    st.info("ℹ️ Aucun fichier de comparaison n'a été téléversé (facultatif).")



# ----------------------------- Analyse simple -----------------------------



def analyse_simplifiee(df):
    st.subheader("Présentes vs Manquantes – Méthode simple")
    total = len(df)
    resume = []

    for col in df.columns:
        if col.lower() in ['timestamp', 'notes']:
            continue

        s = series_with_true_nans(df[col])
        presente = s.notna().sum()
        manquantes = total - presente
        pct_presente = 100 * presente / total if total > 0 else 0
        pct_manquantes = 100 - pct_presente
        statut = "🟢" if pct_presente >= 80 else ("🟠" if pct_presente > 0 else "🔴")

        resume.append({
            "Capteur": str(col).strip(),
            "Présentes": int(presente),
            "% Présentes": round(pct_presente, 2),
            "Manquantes": int(manquantes),
            "% Manquantes": round(pct_manquantes, 2),
            "Statut": statut
        })

    df_resume = pd.DataFrame(resume)
    df_resume["Nom_nettoye"] = df_resume["Capteur"].astype(str).apply(nettoyer_nom_capteur)

    st.dataframe(df_resume, use_container_width=True)
    return df_resume



df_simple = analyse_simplifiee(df_main)



df_simple["Capteur"] = df_simple["Capteur"].astype(str).str.strip()
df_simple["Doublon"] = df_simple["Capteur"].duplicated(keep=False) \
    .map({True: "🔁 Oui", False: "✅ Non"})
df_simple = df_simple.drop_duplicates(subset=["Nom_nettoye"], keep="last").reset_index(drop=True)



if capteurs_reference_cleaned and len(capteurs_reference_cleaned) > 0:
    df_simple["Dans la référence"] = df_simple["Nom_nettoye"].isin(capteurs_reference_cleaned) \
        .map({True: "✅ Oui", False: "❌ Non"})
    df_simple = df_simple.sort_values(by="Dans la référence", ascending=False).reset_index(drop=True)

    st.subheader("✅ Capteurs trouvés dans la référence")
    df_valides = df_simple[df_simple["Dans la référence"] == "✅ Oui"]
    if not df_valides.empty:
        st.dataframe(df_valides[["Capteur", "Dans la référence", "Doublon"]], use_container_width=True)
    else:
        st.markdown("Aucun capteur valide trouvé.")

    st.subheader("❌ Capteurs absents de la référence")
    df_non_valides = df_simple[df_simple["Dans la référence"] == "❌ Non"]
    if not df_non_valides.empty:
        st.dataframe(df_non_valides[["Capteur", "Dans la référence", "Doublon"]], use_container_width=True)
    else:
        st.markdown("Tous les capteurs sont présents dans la référence.")

    if not df_non_valides.empty:
        st.subheader("Liste brute – Capteurs du fichier principal absents de la référence")
        st.write(df_non_valides["Capteur"].tolist())

    capteurs_trouves = set(df_simple["Nom_nettoye"])
    manquants = sorted(capteurs_reference_cleaned - capteurs_trouves)
    if manquants:
        st.subheader("Capteurs attendus non trouvés dans les données analysées")
        st.markdown("Voici les capteurs présents dans le fichier de référence mais absents du fichier principal :")
        df_manquants = pd.DataFrame(manquants, columns=["Capteur (référence manquant dans les données)"])
        st.dataframe(df_manquants, use_container_width=True)
    else:
        st.markdown("✅ Tous les capteurs attendus sont présents dans les données.")



# ----------------------------- Resample -----------------------------



def resampler_df(df, frequence_str):
    if "timestamp" not in df.columns:
        st.warning("⚠️ Colonne 'timestamp' non trouvée dans le fichier.")
        return df

    st.info(f"⏱️ Fréquence sélectionnée : {frequence_str}")

    if frequence_str == "1min":
        st.info("✅ Pas de rééchantillonnage nécessaire (1min).")
        return df.copy()

    try:
        df = df.copy()
        df["timestamp"] = pd.to_datetime(df["timestamp"], errors="coerce")
        df = df.dropna(subset=["timestamp"])
        freq = rule_map[frequence_str]
        df["timestamp"] = df["timestamp"].dt.floor(freq)
        df = df.set_index("timestamp")

        agg_map = {}
        for col in df.columns:
            if col.lower() in ("notes",):
                agg_map[col] = "first"
            else:
                agg_map[col] = "mean" if is_numeric_dtype(df[col]) else "first"

        df_resampled = df.resample(rule_map[frequence_str]).agg(agg_map).reset_index()
        st.success(f"✅ Données rééchantillonnées avec succès à {frequence_str}.")
        return df_resampled

    except Exception as e:
        st.error(f"❌ Erreur lors du rééchantillonnage : {e}")
        return df.reset_index()



# ----------------------------- Nouvelle complétude (INDICATEUR) -----------------------------



def build_presence_indicator(df: pd.DataFrame) -> pd.DataFrame:
    cols = [c for c in df.columns if str(c).lower() not in ("timestamp", "notes")]
    ind = pd.DataFrame(index=df.index)
    for c in cols:
        s = series_with_true_nans(df[c])
        ind[c] = s.notna().astype("float")
    return ind



def analyser_completude_freq(df: pd.DataFrame, frequence_str: str, rule_map: dict) -> pd.DataFrame:
    if "timestamp" not in df.columns:
        st.error("❌ La colonne 'timestamp' est manquante.")
        return pd.DataFrame()

    df = df.copy()
    df["timestamp"] = pd.to_datetime(df["timestamp"], errors="coerce")
    df = df.dropna(subset=["timestamp"]).sort_values("timestamp").reset_index(drop=True)

    ind = build_presence_indicator(df)
    ind.index = df["timestamp"]

    if frequence_str != "1min":
        freq = rule_map[frequence_str]
        ind_bin = ind.resample(freq).max()
        total_expected = float(len(ind_bin.index))
        total_present = ind_bin.fillna(0).sum(axis=0)
    else:
        total_expected = float(len(ind))
        total_present = ind.sum(axis=0)

    rows = []
    for c in ind.columns:
        pres = float(total_present.get(c, 0.0))
        pct = 100.0 * pres / total_expected if total_expected > 0 else 0.0
        statut = "🟢" if pct >= 80 else ("🟠" if pct > 0 else "🔴")
        rows.append({
            "Capteur": c,
            "Présentes": int(round(pres)),
            "% Présentes": round(pct, 2),
            "Manquantes": int(round(total_expected - pres)),
            "% Manquantes": round(100.0 - pct, 2),
            "Statut": statut
        })

    return pd.DataFrame(rows)



# ----------------------------- Analyse de complétude (fiable) -----------------------------



st.subheader(f"📈 Analyse de complétude des données brutes ({frequence})")

stats_main = analyser_completude_freq(df_main, frequence, rule_map)
stats_main = stats_main.drop_duplicates(subset=["Capteur"], keep="last").reset_index(drop=True)

st.dataframe(stats_main, use_container_width=True)



# ----------------------------- 🆕 Statistiques descriptives -----------------------------



st.subheader("📐 Statistiques descriptives par capteur (Min / Max / Moyenne / Médiane / Écart-type)")

df_stats = calculer_stats_descriptives(df_main)
df_stats["Nom_nettoye"] = df_stats["Capteur"].apply(nettoyer_nom_capteur)
df_stats = df_stats.drop_duplicates(subset=["Nom_nettoye"], keep="last").drop(columns=["Nom_nettoye"]).reset_index(drop=True)

# Fusionner avec les statuts de complétude pour contexte
df_stats_merged = df_stats.merge(
    stats_main[["Capteur", "Statut", "% Présentes"]],
    on="Capteur",
    how="left"
)

# Affichage avec mise en forme conditionnelle via styler
def colorier_statut(val):
    if val == "🟢":
        return "background-color: #C6EFCE; color: #006100"
    elif val == "🟠":
        return "background-color: #FFEB9C; color: #9C5700"
    elif val == "🔴":
        return "background-color: #FFC7CE; color: #9C0006"
    return ""

try:
    styled = df_stats_merged.style.map(colorier_statut, subset=["Statut"])
except AttributeError:
    styled = df_stats_merged.style.applymap(colorier_statut, subset=["Statut"])
st.dataframe(styled, use_container_width=True)

# Petit résumé rapide sous le tableau
col1, col2, col3 = st.columns(3)
numeric_cols = df_stats[df_stats["N valeurs"] > 0]
if not numeric_cols.empty:
    with col1:
        capteur_max = numeric_cols.loc[numeric_cols["Max"].idxmax(), "Capteur"]
        val_max = numeric_cols["Max"].max()
        st.metric("Valeur max globale", f"{val_max:.4g}", help=f"Capteur : {capteur_max}")
    with col2:
        capteur_min = numeric_cols.loc[numeric_cols["Min"].idxmin(), "Capteur"]
        val_min = numeric_cols["Min"].min()
        st.metric("Valeur min globale", f"{val_min:.4g}", help=f"Capteur : {capteur_min}")
    with col3:
        st.metric("Capteurs avec données numériques", f"{len(numeric_cols)}")



# ----------------------------- Légende + Résumé -----------------------------



st.markdown("""
### 🧾 Légende des statuts :
- 🟢 : Capteur exploitable (≥80 % de valeurs présentes)
- 🟠 : Incomplet (entre 1 % et 79 %)
- 🔴 : Données absentes (0 %)
""")

count_vert = stats_main["Statut"].value_counts().get("🟢", 0)
count_orange = stats_main["Statut"].value_counts().get("🟠", 0)
count_rouge = stats_main["Statut"].value_counts().get("🔴", 0)

st.markdown(f"""
**Résumé des capteurs :**
- Capteurs exploitables (🟢) : `{count_vert}`
- Capteurs incomplets (🟠) : `{count_orange}`
- Capteurs vides (🔴) : `{count_rouge}`
""")

st.caption(f"⏱️ Lignes analysées : {len(df_main)}")
st.caption(
    f"🧮 Colonnes (hors timestamp/notes) : "
    f"{len([c for c in df_main.columns if str(c).lower() not in ('timestamp','notes')])}"
)



# ----------------------------- Graphique complétude -----------------------------

df_plot = stats_main.sort_values(by="% Présentes", ascending=True)
fig, ax = plt.subplots(figsize=(10, max(6, len(df_plot) * 0.25)))
sns.barplot(
    data=df_plot,
    y="Capteur",
    x="% Présentes",
    hue="Statut",
    dodge=False,
    palette={"🟢": "green", "🟠": "orange", "🔴": "red"},
    ax=ax
)
plt.title("Complétude des capteurs", fontsize=14)
plt.xlabel("% Données présentes")
plt.ylabel("Capteur")
plt.xlim(0, 100)
plt.tight_layout()
st.pyplot(fig)



# ----------------------------- 🆕 Graphique : distribution des moyennes -----------------------------

#st.subheader("📊 Valeurs moyennes par capteur")

#st.caption(
   # "Ce graphique montre la **valeur moyenne mesurée** par chaque capteur sur toute la période analysée. "
   # "La barre représente la moyenne (ex: 42 kW en moyenne pour un compresseur). "
  #  "Les petites lignes noires aux extrémités (barres d'erreur) indiquent l'écart-type : "
 #   "plus elles sont longues, plus la valeur oscille beaucoup autour de la moyenne. "
 #   "La couleur suit le statut de complétude (🟢 bleu = données fiables, 🟠 orange = données incomplètes, 🔴 rouge = absent)."
#)

#df_moy = df_stats_merged.dropna(subset=["Moyenne"]).sort_values("Moyenne", ascending=True)

#if not df_moy.empty:
 #   fig2, ax2 = plt.subplots(figsize=(10, max(6, len(df_moy) * 0.25)))
 #   palette_moy = {"🟢": "steelblue", "🟠": "orange", "🔴": "red"}
 #   couleurs = [palette_moy.get(s, "gray") for s in df_moy["Statut"]]
 #   ax2.barh(df_moy["Capteur"], df_moy["Moyenne"], color=couleurs)
 #   ax2.errorbar(
 #       df_moy["Moyenne"],
 #       range(len(df_moy)),
 #       xerr=df_moy["Écart-type"].fillna(0),
 #       fmt="none",
 #       color="black",
 #       alpha=0.5,
 #      linewidth=1,
 #       capsize=3,
  #      label="± Écart-type"
  #  )
  #  ax2.set_xlabel("Valeur moyenne (unité selon le capteur)")
  #  ax2.set_ylabel("Capteur")
  #  ax2.set_title("Valeur moyenne par capteur (barres d'erreur = ±1 écart-type)")
  #  ax2.legend()
  #  plt.tight_layout()
  #  st.pyplot(fig2)
#else:
  #  st.info("Aucune colonne numérique disponible pour ce graphique.")



# ----------------------------- Export Excel (style rapport de référence) -----------------------------



import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter



def _border_thin():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def _border_medium_bottom():
    thin = Side(style="thin", color="BFBFBF")
    med  = Side(style="medium", color="1F4E79")
    return Border(left=thin, right=thin, top=thin, bottom=med)


# Palettes couleurs identiques au rapport de référence
COLOR_TITRE_BG    = "0D1F3C"   # bleu très foncé  — ligne titre principale
COLOR_SOUS_BG     = "2E4057"   # bleu foncé        — ligne sous-titre
COLOR_SECTION_BG  = "1F3864"   # bleu section      — séparateur de groupe
COLOR_HEADER_BG   = "1F4E79"   # bleu en-têtes colonnes
COLOR_WHITE       = "FFFFFF"
COLOR_ROW_EVEN    = "F2F2F2"   # gris très clair
COLOR_ROW_ODD     = "FFFFFF"   # blanc
COLOR_VERT_BG     = "C6EFCE";  COLOR_VERT_FG    = "006100"
COLOR_ORANGE_BG   = "FFC000";  COLOR_ORANGE_FG  = "7F4800"
COLOR_ROUGE_BG    = "FFC7CE";  COLOR_ROUGE_FG   = "9C0006"

FONT_NAME = "Arial"


def style_titre(ws, row, text, n_cols, bg=COLOR_TITRE_BG, size=14):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(name=FONT_NAME, bold=True, size=size, color=COLOR_WHITE)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
    ws.row_dimensions[row].height = 22


def style_sous_titre(ws, row, text, n_cols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(name=FONT_NAME, size=10, color=COLOR_WHITE)
    cell.fill = PatternFill("solid", fgColor=COLOR_SOUS_BG)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 16


def style_section(ws, row, text, n_cols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(name=FONT_NAME, bold=True, size=11, color=COLOR_WHITE)
    cell.fill = PatternFill("solid", fgColor=COLOR_SECTION_BG)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 18


def style_headers(ws, row, headers):
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = Font(name=FONT_NAME, bold=True, size=10, color=COLOR_WHITE)
        cell.fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _border_medium_bottom()
    ws.row_dimensions[row].height = 30


def style_data_row(ws, row_idx, data_row, statut_col_idx=None):
    """Écrit une ligne de données avec alternance de couleurs + colorisation statut."""
    is_even = (row_idx % 2 == 0)
    base_bg = COLOR_ROW_EVEN if is_even else COLOR_ROW_ODD

    for col_idx, val in enumerate(data_row, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.font = Font(name=FONT_NAME, size=10)
        cell.border = _border_thin()

        # Colorisation spéciale colonne Statut
        if statut_col_idx and col_idx == statut_col_idx:
            if val == "🟢":
                cell.fill = PatternFill("solid", fgColor=COLOR_VERT_BG)
                cell.font = Font(name=FONT_NAME, size=10, color=COLOR_VERT_FG, bold=True)
            elif val == "🟠":
                cell.fill = PatternFill("solid", fgColor=COLOR_ORANGE_BG)
                cell.font = Font(name=FONT_NAME, size=10, color=COLOR_ORANGE_FG, bold=True)
            elif val == "🔴":
                cell.fill = PatternFill("solid", fgColor=COLOR_ROUGE_BG)
                cell.font = Font(name=FONT_NAME, size=10, color=COLOR_ROUGE_FG, bold=True)
            else:
                cell.fill = PatternFill("solid", fgColor=base_bg)
        else:
            cell.fill = PatternFill("solid", fgColor=base_bg)

        cell.alignment = Alignment(horizontal="left" if col_idx == 1 else "center", vertical="center", wrap_text=False)


def auto_col_width(ws, col_idx, df_col_values, header, min_w=10, max_w=60):
    max_len = max(len(str(header)), max(
        (len(str(v)) for v in df_col_values if v is not None), default=0
    ))
    ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, min_w), max_w)


def ecrire_feuille_style(wb, sheet_name, titre, sous_titre, df, statut_col_name=None, section_label=None):
    """Crée une feuille formatée dans le style du rapport de référence."""
    ws = wb.create_sheet(title=sheet_name)
    n_cols = len(df.columns)
    current_row = 1

    # Titre + sous-titre
    style_titre(ws, current_row, titre, n_cols)
    current_row += 1
    style_sous_titre(ws, current_row, sous_titre, n_cols)
    current_row += 1

    # Ligne vide
    current_row += 1

    # Section optionnelle
    if section_label:
        style_section(ws, current_row, section_label, n_cols)
        current_row += 1

    # En-têtes colonnes
    headers = list(df.columns)
    style_headers(ws, current_row, headers)
    header_row = current_row
    current_row += 1

    # Trouver l'index de la colonne Statut (1-based)
    statut_col_idx = None
    if statut_col_name and statut_col_name in df.columns:
        statut_col_idx = df.columns.get_loc(statut_col_name) + 1

    # Données
    for _, row_data in df.iterrows():
        style_data_row(ws, current_row, list(row_data), statut_col_idx=statut_col_idx)
        current_row += 1

    # Largeur des colonnes
    for col_idx, col_name in enumerate(df.columns, start=1):
        auto_col_width(ws, col_idx, df[col_name].tolist(), col_name)

    # Figer la ligne d'en-têtes
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    return ws


def generer_rapport_excel(df_simple, stats_main, df_stats_merged,
                           df_non_valides=None, df_manquants=None,
                           periode="", frequence=""):
    wb = Workbook()
    # Supprimer la feuille par défaut
    wb.remove(wb.active)

    sous_titre_base = f"Période : {periode}  |  Fréquence : {frequence}"

    # --- Feuille 1 : Complétude brute ---
    ecrire_feuille_style(
        wb,
        sheet_name="Complétude brute",
        titre="RAPPORT D'ANALYSE — COMPLÉTUDE DES CAPTEURS",
        sous_titre=sous_titre_base,
        df=stats_main,
        statut_col_name="Statut",
        section_label="📈  ANALYSE DE COMPLÉTUDE — Données brutes"
    )

    # --- Feuille 2 : Stats descriptives ---
    ecrire_feuille_style(
        wb,
        sheet_name="Stats descriptives",
        titre="RAPPORT D'ANALYSE — STATISTIQUES DESCRIPTIVES",
        sous_titre=sous_titre_base,
        df=df_stats_merged,
        statut_col_name="Statut",
        section_label="📐  MIN / MAX / MOYENNE / MÉDIANE / ÉCART-TYPE par capteur"
    )

    # --- Feuille 3 : Résumé capteurs ---
    ecrire_feuille_style(
        wb,
        sheet_name="Résumé capteurs",
        titre="RÉSUMÉ DES CAPTEURS",
        sous_titre=sous_titre_base,
        df=df_simple,
        statut_col_name="Statut",
        section_label="📋  DISPONIBILITÉ COMPLÈTE DES DONNÉES PAR CAPTEUR"
    )

    # --- Feuille 4 : Capteurs non reconnus (facultatif) ---
    if df_non_valides is not None and not df_non_valides.empty:
        ecrire_feuille_style(
            wb,
            sheet_name="Capteurs non reconnus",
            titre="CAPTEURS ABSENTS DE LA RÉFÉRENCE",
            sous_titre=sous_titre_base,
            df=df_non_valides,
            statut_col_name="Dans la référence"
        )

    # --- Feuille 5 : Capteurs manquants (facultatif) ---
    if df_manquants is not None and not df_manquants.empty:
        ecrire_feuille_style(
            wb,
            sheet_name="Capteurs manquants",
            titre="CAPTEURS ATTENDUS NON TROUVÉS",
            sous_titre=sous_titre_base,
            df=df_manquants
        )

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# --- Génération et bouton téléchargement ---

st.subheader("📤 Export des résultats (Excel)")

periode_str = ""
if not df_main.empty:
    periode_str = f"{df_main['timestamp'].min().strftime('%d %b %Y')} – {df_main['timestamp'].max().strftime('%d %b %Y')}"

rapport_bytes = generer_rapport_excel(
    df_simple=df_simple,
    stats_main=stats_main,
    df_stats_merged=df_stats_merged,
    df_non_valides=df_non_valides if 'df_non_valides' in locals() else None,
    df_manquants=df_manquants if 'df_manquants' in locals() else None,
    periode=periode_str,
    frequence=frequence
)

st.download_button(
    label="📥 Télécharger le rapport Excel",
    data=rapport_bytes,
    file_name="rapport_capteurs.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
